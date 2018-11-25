from openpyxl import Workbook
from openpyxl import load_workbook
import sys

wb = load_workbook('./sample.xlsx')
ws = wb['platform']
print(ws['A1'].value)
print(wb.sheetnames)
