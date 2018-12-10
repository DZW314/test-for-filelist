from openpyxl import Workbook
from openpyxl import load_workbook
from zlib import crc32
import sys
import glob
import logging
import xml.etree.ElementTree as ET
def GetCrc32(filename): # calculate crc32
    with open(filename, 'rb') as f:
        return crc32(f.read())

def strnset(str,ch,n): # string change
    str = str[:n] + ch
    return str

#log setting
LOG_FORMAT = "%(asctime)s - %(levelname)s - %(message)s"
logging.basicConfig(filename='info.log', level=logging.DEBUG, format=LOG_FORMAT)

# if len(sys.argv) < 2:
#     print('You must enter the file')
#     exit(1)
# elif len(sys.argv) > 2:
#     print('Only one file is permitted')
#     exit(1)

#filename = sys.argv[1]
#search all files and store in an array
logging.info("Start search all files in Packages.")
listFile = []
listuxz = glob.glob("./*/*.uxz")
listbin = glob.glob("./*/*.bin")
listtgz = glob.glob("./*/*.tgz")
listexe = glob.glob("./*/*.exe")
listFile.extend(listuxz)
listFile.extend(listbin)
listFile.extend(listtgz)
listFile.extend(listexe)

logging.info("Finish searching.")
#list sort
logging.info("Sorting the list...")
sortFile = sorted(listFile)

#Create sheet for excel
wb = Workbook()
ws = wb.active
ws.title = "platform"
ws1 = wb.create_sheet("win2019")
ws2 = wb.create_sheet("win2016")
ws3 = wb.create_sheet("rhel7")
ws4 = wb.create_sheet("suse12")

#Check all packages by crc32
row = 1
for n in sortFile:
    crcpkg = format(GetCrc32(n), 'x')
    print('{:s} {:8} {:s}'.format( n,' crc32: ', crcpkg))
    logging.info('{:s} {:8} {:s}'.format( n,' crc32: ', crcpkg))

    tmpxml = strnset(n,".xml",-4)
    tree = ET.parse(tmpxml)
    # print(tree.getroot())
    root = tree.getroot()
    crcxml = root.findall(".//*[@NAME='crc']/VALUE")
    for tmp in crcxml:
        # print(tmp.text)
        ws.cell(column=4, row=row, value=tmp.text)
    ws.cell(column=1, row=row, value=n)
    ws.cell(column=2, row=row, value=crcpkg)
    ws.cell(column=3, row=row, value=tmpxml)

    row = row + 1

wb.save("sample.xlsx")
logging.info("Mission Completed!")
# result =  'crc32.txt'
# f = open ('./' + result,'w')
# for n in sortFile:
#     str = n;
#     crc = format(getCrc32(n), 'x')
#     print('{:s} {:8} {:x}'.format( n,' crc32: ', getCrc32(n)))
#     f.write(str + ' page_crc32: ' + crc + '\n')
# f.close()
