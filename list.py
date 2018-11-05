import os
from  openpyxl import Workbook
import datetime
system = os.name
env = os.environ
filePath = os.path.abspath('.')
print(system)
print(filePath)

wb = Workbook()
ws = wb.active
ws['A1'] = 42
ws.append([1,2,3])
ws['b1'] = datetime.datetime.now()
ws['A3'] = system
ws['A4'] = filePath
wb.save("sample.xlsx")
